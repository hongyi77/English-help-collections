import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4.title = "CET4常用单词"

wb6 = openpyxl.Workbook()
ws6 = wb6.active
ws6.title = "CET6常用单词"

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["序号", "单词", "音标", "词性", "中文释义", "例句"]

def write_header(ws):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

write_header(ws4)
write_header(ws6)

cet4 = [
[1,"abandon","/əˈbændən/","v.","放弃；遗弃","He had to abandon his car in the snow and walk home."],
[2,"ability","/əˈbɪləti/","n.","能力；才能","She has the ability to solve complex problems quickly."],
[3,"abnormal","/æbˈnɔːrml/","adj.","反常的；变态的","The doctor noticed abnormal cells in the test results."],
[4,"above","/əˈbʌv/","adv.","在上面；上述的","The temperature rose well above freezing."],
[5,"abroad","/əˈbrɔːd/","adv.","在国外；到国外","Many students go abroad to pursue higher education."],
[6,"absorb","/əbˈzɔːrb/","v.","吸收；吸引注意力","Plants absorb carbon dioxide from the atmosphere."],
[7,"abstract","/ˈæbstrækt/","adj.","抽象的；理论的","The concept of justice is quite abstract."],
[8,"academic","/ˌækəˈdemɪk/","adj.","学术的；学院的","She has an excellent academic record."],
[9,"accelerate","/əkˈseləreɪt/","v.","加速；促进","The car accelerated rapidly on the highway."],
[10,"access","/ˈækses/","n./v.","通道；接近；获取","Students need access to reliable research materials."],
[11,"accident","/ˈæksɪdənt/","n.","事故；意外","He was injured in a car accident last week."],
[12,"accommodate","/əˈkɑːmədeɪt/","v.","容纳；适应；提供住宿","The hotel can accommodate up to 500 guests."],
[13,"accompany","/əˈkʌmpəni/","v.","陪伴；伴随","She accompanied her friend to the hospital."],
[14,"accomplish","/əˈkɑːmplɪʃ/","v.","完成；实现","They accomplished the mission ahead of schedule."],
[15,"according to","/əˈkɔːrdɪŋ tuː/","prep.","根据；按照","According to the latest survey, most people prefer online shopping."],
[16,"account","/əˈkaʊnt/","n./v.","账户；解释；描述","He gave a detailed account of what happened."],
[17,"accumulate","/əˈkjuːmjəleɪt/","v.","积累；积聚","Dust tends to accumulate if you do not clean regularly."],
[18,"accurate","/ˈækjərət/","adj.","准确的；精确的","The scientist made accurate measurements of the sample."],
[19,"achieve","/əˈtʃiːv/","v.","达到；取得；实现","She achieved her goal of becoming a doctor."],
[20,"acknowledge","/əkˈnɑːlɪdʒ/","v.","承认；致谢；认可","He acknowledged his mistake publicly."],
[21,"acquire","/əˈkwaɪər/","v.","获得；学到","She acquired a good knowledge of English through practice."],
[22,"adapt","/əˈdæpt/","v.","适应；改编","It took time to adapt to the new environment."],
[23,"adequate","/ˈædɪkwət/","adj.","充足的；适当的","We need adequate funding to complete the project."],
[24,"adjust","/əˈdʒʌst/","v.","调整；适应","You need to adjust the mirror before driving."],
[25,"administration","/ədˌmɪnɪˈstreɪʃn/","n.","管理；行政","The administration announced a new policy on education."],
[26,"admire","/ədˈmaɪər/","v.","钦佩；赞赏","I admire her courage and determination."],
[27,"admission","/ədˈmɪʃn/","n.","允许进入；承认；入场费","Admission to the museum is free on Sundays."],
[28,"adopt","/əˈdɑːpt/","v.","采纳；收养；采用","The company adopted a new marketing strategy."],
[29,"advance","/ədˈvæns/","v./n.","前进；进步；预付款","Technology has advanced rapidly in recent years."],
[30,"advantage","/ədˈvæntɪdʒ/","n.","优势；有利条件","Her experience gave her a significant advantage."],
[31,"adventure","/ədˈventʃər/","n.","冒险；奇遇","The trip was an exciting adventure."],
[32,"advertise","/ˈædvərtaɪz/","v.","做广告；宣传","The company advertised the product on television."],
[33,"advisable","/ədˈvaɪzəbl/","adj.","明智的；可取的","It is advisable to arrive early for the interview."],
[34,"advocate","/ˈædvəkeɪt/","v./n.","提倡；拥护者","She advocates for equal rights for all people."],
[35,"affair","/əˈfer/","n.","事情；事件；事务","The government is investigating the scandalous affair."],
[36,"affect","/əˈfekt/","v.","影响；感动","The weather can affect your mood significantly."],
[37,"affiliate","/əˈfɪlieɪt/","v./n.","使隶属；附属机构","The company is affiliated with a larger corporation."],
[38,"afford","/əˈfɔːrd/","v.","负担得起；提供","They cannot afford to buy a new house right now."],
[39,"aggressive","/əˈɡresɪv/","adj.","好斗的；有进取心的","He has an aggressive approach to business."],
[40,"agreement","/əˈɡriːmənt/","n.","协议；同意","The two countries signed a trade agreement."],
[41,"agriculture","/ˈæɡrɪkʌltʃər/","n.","农业","Agriculture remains important in rural areas."],
[42,"alert","/əˈlɜːrt/","adj./n.","警觉的；警报","The fire alert warned residents to evacuate immediately."],
[43,"allocate","/ˈæləkeɪt/","v.","分配；拨出","The government allocated funds for disaster relief."],
[44,"allowance","/əˈlaʊəns/","n.","津贴；零花钱","He receives a monthly travel allowance from his employer."],
[45,"alter","/ˈɔːltər/","v.","改变；修改","She had to alter her plans due to bad weather."],
[46,"alternative","/ɔːlˈtɜːrnətɪv/","n./adj.","替代方案；可选择的","We need to find an alternative solution to this problem."],
[47,"amaze","/əˈmeɪz/","v.","使惊讶","Her performance amazed everyone in the audience."],
[48,"ambition","/æmˈbɪʃn/","n.","雄心；抱负","His ambition is to become a successful entrepreneur."],
[49,"ambitious","/æmˈbɪʃəs/","adj.","有雄心的；野心勃勃的","She is an ambitious student who always aims for excellence."],
[50,"amount","/əˈmaʊnt/","n./v.","数量；总计；等于","A large amount of money was spent on the project."],
[51,"analyze","/ˈænəlaɪz/","v.","分析；解析","Scientists analyze data to draw meaningful conclusions."],
[52,"ancient","/ˈeɪnʃənt/","adj.","古代的；古老的","The ancient city has attracted millions of tourists."],
[53,"annual","/ˈænjuəl/","adj./n.","每年的；年度刊物","The company publishes an annual report on its finances."],
[54,"anticipate","/ænˈtɪsɪpeɪt/","v.","预期；期望","We anticipate that sales will increase next quarter."],
[55,"anxiety","/æŋˈzaɪəti/","n.","焦虑；忧虑","Many students experience anxiety before exams."],
[56,"anonymous","/əˈnɑːnɪməs/","adj.","匿名的","The donation was made by an anonymous benefactor."],
[57,"appeal","/əˈpiːl/","v./n.","呼吁；吸引；上诉","The charity appealed for donations to help the poor."],
[58,"appear","/əˈpɪr/","v.","出现；显得","A rainbow appeared after the rain stopped."],
[59,"application","/ˌæplɪˈkeɪʃn/","n.","申请；应用","She submitted her application for the scholarship."],
[60,"appoint","/əˈpɔɪnt/","v.","任命；指定","They appointed a new manager to lead the department."],
[61,"appreciate","/əˈpriːʃieɪt/","v.","欣赏；感激；领会","I really appreciate your help with this project."],
[62,"approach","/əˈproʊtʃ/","v./n.","接近；处理；方法","We need a new approach to solve this issue."],
[63,"appropriate","/əˈproʊpriət/","adj.","适当的；合适的","Please wear appropriate clothing for the interview."],
[64,"approve","/əˈpruːv/","v.","批准；赞成","The committee approved the proposed changes."],
[65,"approximate","/əˈprɑːksɪmət/","adj.","大约的；近似的","The approximate cost of the project is around one million dollars."],
[66,"argue","/ˈɑːrɡjuː/","v.","争论；主张","They argued about the best way to proceed."],
[67,"arise","/əˈraɪz/","v.","出现；产生","New challenges arise with every new opportunity."],
[68,"arrange","/əˈreɪndʒ/","v.","安排；排列","She arranged a meeting for next Monday."],
[69,"artificial","/ˌɑːrtɪˈfɪʃl/","adj.","人工的；人造的","Artificial intelligence is transforming many industries."],
[70,"assess","/əˈses/","v.","评估；评价","Teachers assess student progress through examinations."],
[71,"assign","/əˈsaɪn/","v.","分配；指派","The teacher assigned homework to the students."],
[72,"assist","/əˈsɪst/","v.","帮助；协助","The nurse assisted the doctor during the operation."],
[73,"associate","/əˈsoʊʃieɪt/","v./adj.","联系；联想；联合的","People often associate summer with vacation."],
[74,"assume","/əˈsuːm/","v.","假设；承担","I assume you have already read the report."],
[75,"assure","/əˈʃʊr/","v.","保证；确保","I assure you that everything will be fine."],
[76,"atmosphere","/ˈætməsfɪr/","n.","大气层；气氛","The restaurant had a warm and friendly atmosphere."],
[77,"attach","/əˈtætʃ/","v.","附上；系；依恋","Please attach your resume to the email."],
[78,"attempt","/əˈtempt/","v./n.","尝试；企图","He attempted to climb the mountain alone."],
[79,"attend","/əˈtend/","v.","参加；出席；照顾","Over 200 people attended the conference."],
[80,"attribute","/əˈtrɪbjuːt/","n./v.","属性；归因于","She attributed her success to hard work and perseverance."],
]

cet6 = [
[1,"abolish","/əˈbɑːlɪʃ/","v.","废除；废止","The government abolished the outdated law."],
[2,"abortion","/əˈbɔːrʃn/","n.","流产；堕胎","The debate on abortion continues to divide public opinion."],
[3,"abrupt","/əˈbrʌpt/","adj.","突然的；唐突的","The bus came to an abrupt halt in the storm."],
[4,"absurd","/əbˈsɜːrd/","adj.","荒谬的；可笑的","It is absurd to blame him for something he did not do."],
[5,"abundance","/əˈbʌndəns/","n.","丰富；充裕","There is an abundance of fresh produce at the market."],
[6,"accessible","/əkˈsesəbl/","adj.","可接近的；可使用的","The building is accessible to wheelchair users."],
[7,"acquisition","/ˌækwɪˈzɪʃn/","n.","获得；收购","The company announced the acquisition of a smaller rival."],
[8,"adapt","/əˈdæpt/","v.","适应；改编","Animals must adapt to survive in changing environments."],
[9,"addiction","/əˈdɪkʃn/","n.","上瘾；沉迷","Internet addiction is becoming a serious social problem."],
[10,"adolescent","/ˌædəˈlesnt/","n./adj.","青少年；青春期的","Adolescents face unique challenges during puberty."],
[11,"aesthetic","/esˈθetɪk/","adj.","审美的；美学的","The building has great aesthetic appeal."],
[12,"afflict","/əˈflɪkt/","v.","折磨；使痛苦","The disease afflicts millions of people worldwide."],
[13,"aggravate","/ˈæɡrəveɪt/","v.","加重；恶化；激怒","Lack of sleep can aggravate health problems."],
[14,"aggregate","/ˈæɡrɪɡət/","n./adj.","总计；合计的","The aggregate score determined the competition winner."],
[15,"allegation","/ˌæləˈɡeɪʃn/","n.","指控；断言","The politician denied all allegations of corruption."],
[16,"alliance","/əˈlaɪəns/","n.","联盟；联合","The two countries formed a military alliance."],
[17,"allocate","/ˈæləkeɪt/","v.","分配；拨出","Resources must be allocated efficiently."],
[18,"alleviate","/əˈliːvieɪt/","v.","减轻；缓和","Medicine can alleviate the symptoms of the illness."],
[19,"allusion","/əˈluːʒn/","n.","暗示；典故","The poem contains allusions to classical mythology."],
[20,"ambiguity","/ˌæmbɪˈɡjuːəti/","n.","模糊；歧义","The ambiguity of the sentence caused confusion."],
[21,"ample","/ˈæmpl/","adj.","充足的；丰富的","There is ample evidence to support the theory."],
[22,"analogy","/əˈnælədʒi/","n.","类比；类推","He drew an analogy between the brain and a computer."],
[23,"anonymous","/əˈnɑːnɪməs/","adj.","匿名的","An anonymous caller reported the suspicious activity."],
[24,"antagonist","/ænˈtæɡənɪst/","n.","对手；敌手；反派","The antagonist in the novel represents evil forces."],
[25,"anthropology","/ˌænθrəˈpɑːlədʒi/","n.","人类学","She studied anthropology at university."],
[26,"apparatus","/ˌæpəˈreɪtəs/","n.","器械；装置","The laboratory apparatus was expensive to maintain."],
[27,"appease","/əˈpiːz/","v.","安抚；平息","Nothing could appease the angry customer."],
[28,"appliance","/əˈplaɪəns/","n.","器具；家电","The kitchen is equipped with modern appliances."],
[29,"arbitrary","/ˈɑːrbəteri/","adj.","任意的；武断的","The decision seemed completely arbitrary."],
[30,"architect","/ˈɑːrkɪtekt/","n.","建筑师；缔造者","The architect designed a stunning skyscraper downtown."],
[31,"articulate","/ɑːrˈtɪkjəleɪt/","v./adj.","清楚表达；口齿清晰的","She articulated her ideas with clarity and precision."],
[32,"aspiration","/ˌæspəˈreɪʃn/","n.","抱负；渴望","He has strong aspirations for a career in medicine."],
[33,"assertion","/əˈsɜːrʃn/","n.","断言；主张","His assertion was not supported by any evidence."],
[34,"assimilate","/əˈsɪmɪleɪt/","v.","吸收；同化","Immigrants often take time to assimilate into the culture."],
[35,"atrocity","/əˈtɑːrəti/","n.","暴行；残忍","The war atrocities shocked the international community."],
[36,"authentic","/ɔːˈθentɪk/","adj.","真实的；可靠的","The museum displayed authentic artifacts from ancient times."],
[37,"autonomous","/ɔːˈtɑːnəməs/","adj.","自治的；自主的","The region became an autonomous territory."],
[38,"avert","/əˈvɜːrt/","v.","避免；转移","Quick thinking helped avert a major disaster."],
[39,"barren","/ˈbærən/","adj.","贫瘠的；不毛的","The barren land could not support any crops."],
[40,"benevolent","/bəˈnevələnt/","adj.","仁慈的；慈善的","The benevolent donor funded scholarships for many students."],
[41,"bewilder","/bɪˈwɪldər/","v.","使迷惑；使不知所措","The complicated instructions bewildered the customers."],
[42,"breach","/briːtʃ/","n./v.","违反；缺口","The breach of contract led to a lawsuit."],
[43,"burgeon","/ˈbɜːrdʒən/","v.","迅速发展；萌芽","The tech industry continues to burgeon in the region."],
[44,"buttress","/ˈbʌtrɪs/","v./n.","支持；扶壁","New evidence buttressed the scientists hypothesis."],
[45,"catastrophe","/kəˈtæstrəfi/","n.","灾难；大祸","The earthquake was a catastrophic event for the city."],
[46,"caution","/ˈkɔːʃn/","n./v.","谨慎；警告","Exercise caution when handling chemicals."],
[47,"chronicle","/ˈkrɑːnɪkl/","n./v.","编年史；记录","The book chronicles the history of the empire."],
[48,"circumscribe","/ˈsɜːrkəmskraɪb/","v.","限制；约束","His authority was circumscribed by the boards rules."],
[49,"coherent","/koʊˈhɪrənt/","adj.","连贯的；一致的","She presented a coherent argument that convinced everyone."],
[50,"collaborate","/kəˈlæbəreɪt/","v.","合作；协作","The two companies collaborated on the research project."],
[51,"commemorate","/kəˈmeməreɪt/","v.","纪念；庆祝","A statue was built to commemorate the hero."],
[52,"compassionate","/kəmˈpæʃənət/","adj.","有同情心的","The compassionate nurse cared for the elderly patients."],
[53,"compelling","/kəmˈpelɪŋ/","adj.","引人注目的；令人信服的","She made a compelling case for the reform."],
[54,"complement","/ˈkɑːmplɪment/","v./n.","补充；互补","The wine complements the meal perfectly."],
[55,"comply","/kəmˈplaɪ/","v.","遵从；服从","All employees must comply with safety regulations."],
[56,"conceive","/kənˈsiːv/","v.","构想；怀孕","She conceived the idea while traveling in Europe."],
[57,"concurrent","/kənˈkʌrənt/","adj.","同时发生的","The concurrent events surprised the investigators."],
[58,"condense","/kənˈdens/","v.","浓缩；压缩","Steam condenses into water droplets on cold surfaces."],
[59,"confine","/kənˈfaɪn/","v.","限制；禁闭","Patients were confined to their rooms during the epidemic."],
[60,"conform","/kənˈfɔːrm/","v.","遵从；符合","Students must conform to the schools dress code."],
[61,"confront","/kənˈfrʌnt/","v.","面对；对抗","We must confront the challenges of climate change."],
[62,"conscience","/ˈkɑːnʃəns/","n.","良心；道德心","His conscience prevented him from telling lies."],
[63,"consensus","/kənˈsensəs/","n.","共识；一致意见","The committee reached a consensus on the proposal."],
[64,"consolidate","/kənˈsɑːlɪdeɪt/","v.","巩固；合并","The company consolidated its operations to reduce costs."],
[65,"constitute","/ˈkɑːnstɪtuːt/","v.","构成；组成","Women constitute over half of the workforce."],
[66,"constrain","/kənˈstreɪn/","v.","限制；约束","Budget constraints limited the projects scope."],
[67,"contemplate","/ˈkɑːntəmpleɪt/","v.","沉思；凝视","She contemplated the meaning of life."],
[68,"contradict","/ˌkɑːntrəˈdɪkt/","v.","反驳；与矛盾","The evidence contradicts his alibi."],
[69,"controversy","/ˈkɑːntrəvɜːrsi/","n.","争议；论战","The policy sparked considerable controversy."],
[70,"convey","/kənˈveɪ/","v.","传达；运输","Words cannot convey my gratitude."],
[71,"corroborate","/kəˈrɑːbəreɪt/","v.","证实；佐证","Witnesses corroborated the suspects story."],
[72,"cosmopolitan","/ˌkɑːzməˈpɑːlɪtn/","adj.","世界主义的；大都市的","New York is a cosmopolitan city with diverse cultures."],
[73,"criterion","/kraɪˈtɪriən/","n.","标准；准则","What criteria do you use to evaluate candidates."],
[74,"crucial","/ˈkruːʃl/","adj.","至关重要的","Education plays a crucial role in development."],
[75,"cultivate","/ˈkʌltɪveɪt/","v.","培养；耕作","She cultivated a passion for reading from a young age."],
[76,"curtail","/kɜːrˈteɪl/","v.","削减；缩减","The government decided to curtail public spending."],
[77,"debilitate","/dɪˈbɪlɪteɪt/","v.","使衰弱；使无力","The illness debilitated him for months."],
[78,"decipher","/dɪˈsaɪfər/","v.","破译；辨认","Archaeologists deciphered the ancient inscription."],
[79,"dedicate","/ˈdedɪkeɪt/","v.","奉献；致力于","She dedicated her life to scientific research."],
[80,"deliberate","/dɪˈlɪbərət/","adj./v.","深思熟虑的；故意地","The decision was deliberate and well thought out."],
]

for i, row in enumerate(cet4, 2):
    for col, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=col, value=val)
        cell.alignment = cell_align
        cell.border = thin_border

for i, row in enumerate(cet6, 2):
    for col, val in enumerate(row, 1):
        cell = ws6.cell(row=i, column=col, value=val)
        cell.alignment = cell_align
        cell.border = thin_border

ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 10
ws4.column_dimensions["E"].width = 25
ws4.column_dimensions["F"].width = 50

ws6.column_dimensions["A"].width = 8
ws6.column_dimensions["B"].width = 20
ws6.column_dimensions["C"].width = 25
ws6.column_dimensions["D"].width = 10
ws6.column_dimensions["E"].width = 25
ws6.column_dimensions["F"].width = 50

path4 = r"D:\英语学习\CET-词汇整理\CET4常用单词.xlsx"
path6 = r"D:\英语学习\CET-词汇整理\CET6常用单词.xlsx"

wb4.save(path4)
wb6.save(path6)

print("CET4 saved: " + path4)
print("CET6 saved: " + path6)
print("Done!")